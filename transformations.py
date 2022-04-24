import openpyxl
from pandas import DataFrame
import datetime


def pr_mon(mon: str) -> str:
    months = {'01': 'январь', '02': 'февраль', '03': 'март', '04': 'апрель', '05': 'май', '06': 'июнь', '07': 'июль',
              '08': 'август', '09': 'сентябрь', '10': 'октябрь', '11': 'ноябрь', '12': 'декабрь'}
    return months[mon]


def to_be(d: int, month: str) -> bool:
    try:
        datetime.date(day=d, month=int(month), year=2022)

    except ValueError:
        return False
    return True


def count_month(date_in: list) -> list:
    mon_lst = []
    for sym in date_in:
        if not (sym.split('.')[1]) in mon_lst:
            mon_lst.append((sym.split('.')[1]))
    return mon_lst


def date_under(day: int, x: str) -> str:
    month = {'01': 'январь', '02': 'февраль', '03': 'март', '04': 'апрель', '05': 'май', '06': 'июнь', '07': 'июль',
             '08': 'август', '09': 'сентябрь', '10': 'октябрь', '11': 'ноябрь', '12': 'декабрь'}
    result_month = ''
    for mon in month.keys():
        if mon == x:
            result_month = mon
    if day < 10:
        day_symbol = "0{}".format(day)
    else:
        day_symbol = str(day)
    return '2022.{}.{}'.format(result_month, day_symbol)


def table_xls(date_start: str, date_end_theory: str, date_start_practice: str, date_end_practice: str,
              consultation: int, exam: int, date_out: list, total: int, time_theory: int, group_data: dict,
              t_k: int, t_e: int, economy_teacher: str, economy: int) -> tuple:
    months = count_month(date_out)
    my_dict = {}
    last_day = date_out[-1]
    for d in range(1, 32):
        my_dict[str(d)] = []
        for y in range(1, len(months) + 1):
            if d == 5:
                my_dict[str(d)].append("Месяц {}".format(pr_mon(months[y - 1])))
            else:
                my_dict[str(d)].append('')
            if to_be(d, months[y - 1]):
                my_dict[str(d)].append(d)
            else:
                my_dict[str(d)].append('')
            if date_start <= date_under(d, months[y - 1]) <= last_day:
                if to_be(d, months[y - 1]):
                    my_dict[str(d)].append(1)
                    my_dict[str(d)].append('В')
                else:
                    my_dict[str(d)].append('')
                    my_dict[str(d)].append("")

            else:
                my_dict[str(d)].append('')
                my_dict[str(d)].append("")
    my_dict['часов'] = ["" for _ in range(1, len(months) * 4 + 1)]
    my_dict['дней'] = ["" for _ in range(1, len(months) * 4 + 1)]
    # print(my_dict)
    for day in date_out:
        day_date = str(int(day.split('.')[2]))
        month_date = day.split('.')[1]
        number = (months.index(month_date) * 4) + 3
        my_dict[day_date][number] = 8
    print(my_dict)
    g = DataFrame(my_dict)
    m = {}
    for i, mon in enumerate(months):
        m[i] = g.loc[((i * 4) + 3)]
    sum_m = []
    for element in m:
        sum_d = sum(sym for sym in m[element] if sym == 8)
        sum_m.append(sum_d)
    for i, mon in enumerate(months):
        my_dict['часов'][(i * 4) + 3] = sum_m[i]
    g = DataFrame(my_dict)
    name = "temp.xlsx"
    g.to_excel(name)
    wb = openpyxl.load_workbook(filename=name)
    sheet = wb["Sheet1"]
    sheet.insert_rows(1, 6)
    sheet.delete_cols(1, 1)
    row = "СРОКИ ОБУЧЕНИЯ С {} ПО {}".format(date_start, exam)
    sheet.merge_cells('A1:AE1')
    sheet.cell(row=1, column=1).value = row
    row = "Учебное заведение ИДПО ФГБОУ ВО Новосибирский ГАУ"
    sheet.merge_cells('A2:AE2')
    sheet.cell(row=2, column=1).value = row
    row = "ПРОФЕССИЯ {} ({}) {} ({})".format(group_data['profession'], group_data['type_preparation'],
                                             group_data['group'], group_data['forma'])
    sheet.merge_cells('A3:AE3')
    sheet.cell(row=3, column=1).value = row
    row = "теория = {} час с {} по {}".format(time_theory, date_start, date_end_theory)
    sheet.cell(row=4, column=1).value = row
    sheet.merge_cells('A4:P4')
    row = "консультация = {} час {}".format(t_k, consultation)
    sheet.cell(row=4, column=17).value = row
    sheet.merge_cells('Q4:AE4')
    row = "практика = {} час с {} по {} ".format((total - time_theory - t_k - t_e), date_start_practice,
                                                 date_end_practice)
    sheet.cell(row=5, column=1).value = row
    sheet.merge_cells('A5:P5')
    row = "квалификационный экзамен = {} час {}".format(t_e, exam)
    sheet.cell(row=5, column=17).value = row
    sheet.merge_cells('Q5:AE5')
    row = "общее число часов  = {} час".format(total)
    sheet.cell(row=6, column=1).value = row
    sheet.merge_cells('A6:P6')
    row = "экономика = {} час {}".format(economy, economy_teacher)
    sheet.cell(row=6, column=17).value = row
    sheet.merge_cells('q6:ae6')
    wb.save("temp.xlsx")
    return name, months
