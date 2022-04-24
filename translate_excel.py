import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import (Border, Side, Alignment, Font)


def translate_file(name_file: str, group_data: dict, months: list) -> None:
    wb = openpyxl.load_workbook(filename=name_file)
    sheet = wb["Sheet1"]
    last_row = 0
    for i in range(1, 32):
        letter = get_column_letter(i)
        sheet.column_dimensions[letter].width = 4

    letter1 = "{}1".format(get_column_letter(1))
    letter2 = "{}2".format(get_column_letter(1))
    letter3 = "{}3".format(get_column_letter(1))
    sheet[letter1].font = Font(bold=True)
    sheet[letter1].alignment = Alignment(horizontal='center')
    sheet[letter2].alignment = Alignment(horizontal='center')
    sheet[letter3].font = Font(bold=True)
    sheet[letter3].alignment = Alignment(horizontal='center')
    for i in range(1, 32):
        letter = "{}7".format(get_column_letter(i))
        sheet[letter].value = ""
    sheet.merge_cells('a7:ae7')
    thins = Side(border_style="thin", color="000000")
    for i in range(1, 32):
        letter = "{}7".format(get_column_letter(i))
        sheet[letter].border = Border(top=thins, bottom=thins, left=thins, right=thins)
    for j in range(len(months)):
        for i in range(1, 34):
            letter = "{}{}".format(get_column_letter(i), j * 4 + 1 + 7)
            sheet[letter].border = Border(top=thins, bottom=thins, left=thins, right=thins)
            letter = "{}{}".format(get_column_letter(i), j * 4 + 1 + 8)
            sheet[letter].border = Border(top=thins, bottom=thins, left=thins, right=thins)
            letter = "{}{}".format(get_column_letter(i), j * 4 + 1 + 9)
            sheet[letter].border = Border(top=thins, bottom=thins, left=thins, right=thins)
            letter = "{}{}".format(get_column_letter(i), j * 4 + 1 + 10)
            sheet[letter].border = Border(top=thins, bottom=thins, left=thins, right=thins)
        letter = "E{}".format(j * 4 + 1 + 7)
        month = sheet[letter].value
        letters = "A{}:AE{}".format((j * 4 + 1 + 7), (j * 4 + 1 + 7))
        sheet.merge_cells(letters)
        letter = "A{}".format(j * 4 + 1 + 7)
        sheet[letter].value = month
        sheet[letter].alignment = Alignment(horizontal='center')
        letter = "AG{}".format(j * 4 + 1 + 9)
        sheet[letter] = "=SUM(A{}:AE{})".format((j * 4 + 1 + 9), (j * 4 + 1 + 9))
        last_row = j * 4 + 1 + 10
    letter = "AF8:AF{}".format(last_row)
    sheet["AF{}".format(last_row + 1)] = "=SUM({})".format(letter)
    letter = "AG8:AG{}".format(last_row)
    sheet["AG{}".format(last_row + 1)] = "=SUM({})".format(letter)
    for i in range(1, 34):
        letter = "{}{}".format(get_column_letter(i), last_row + 1)
        sheet[letter].border = Border(top=thins, bottom=thins, left=thins, right=thins)
    letter = "A{}".format(last_row)
    i = 1
    while sheet[letter].value != None and i < 33:
        i += 1
        letter = "{}{}".format(get_column_letter(i), last_row)
    if i > 2:
        last_cell_1 = "{}{}".format(get_column_letter(i - 1), last_row + 1)
        last_cell_2 = "{}{}".format(get_column_letter(i - 2), last_row + 1)
        sheet[last_cell_2].value = "к"
        sheet[last_cell_2].font = Font(bold=True)
        sheet[last_cell_1].value = "э"
        sheet[last_cell_1].font = Font(bold=True)
    for i in range(1, last_row + 5):
        for j in range(1, 35):
            letter = "{}{}".format(get_column_letter(j), i)
            sheet[letter].font = Font(name='Times New Roman', size=12)
        sheet.row_dimensions[i].height = 15
    letter = "A{}".format(last_row + 4)
    sheet[letter].value = "Преподаватель {}".format(group_data['teacher'])
    sheet[letter].font = Font(name='Times New Roman', size=12)
    letter = "A{}".format(last_row + 5)
    sheet[letter].value = "Количество человек {}".format(group_data['people'])
    sheet[letter].font = Font(name='Times New Roman', size=12)
    sheet['A1'].font = Font(bold=True, name='Times New Roman', size=12)
    sheet['A3'].font = Font(bold=True, name='Times New Roman', size=12)
    name_file_out = "{}-{}-{}.xlsx".format(group_data['profession'], group_data["date_start"], group_data["total_hour"])

    wb.save(name_file_out)
